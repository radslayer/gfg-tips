"""
report_builder.py — the payroll-report calculation + workbook-build logic
for the GFG Payroll admin report page.

v2 (8/31/2026): deductions (PTO / Employee Purchases / Misc Amount /
Misc Reimbursement) now come from live Firestore records instead of an
uploaded Employee Time Off Tracker workbook -- Mike and Thao (manager
role) log these directly in the app, and this report sweeps up every
not-yet-processed record and (when finalize=True, decided by the
caller/main.py) marks it processed with this pay date. See main.py for
the Firestore read/write side; this module only builds the workbook and
summary from data it's handed.

SCOPE NOTE (unchanged from v1, per Rod 8/31/2026): tip allocation uses
ONLY the man-days computed from the single uploaded raw timeclock CSV
for this pay period -- not yet the fully-proven 4-week/2-pay-period
rolling window (that model is proven correct -- see the project doc --
but needs a persisted man-days history across periods that hasn't been
built yet). The report says so explicitly on the Man-Days & Tips Calc
sheet.
"""
import csv
import io
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants (identical to timeclock_to_adp.py)
# ---------------------------------------------------------------------------
BREAKS_TABLE = [
    (0.0, 0),
    (1.833333333, 10),
    (5.666666667, 20),
    (9.5, 30),
    (13.33333333, 40),
    (17.16666667, 50),
]
OT_THRESHOLD_HOURS = 40.0
OT_MULTIPLIER = 1.5
WORKWEEK_START_WEEKDAY = 5  # Saturday

MAN_DAY_FULL_THRESHOLD = 4.0
MAN_DAY_FULL_CREDIT = 1.0
MAN_DAY_HALF_CREDIT = 0.5

CSV_COL_NAME = 7
CSV_COL_DATE = 18
CSV_COL_TRANS_TTL = 25
CSV_DATE_FORMAT = "%d-%b-%Y"

ADP_COLUMNS = [
    "Name", "EE Purchases", "Department", "Regular Hours", "Salary Amount",
    "Overtime Hours", "CC Tips Owed", "NQ Overtime", "NQ Tips", "PTO Hours",
    "Bonus Amount", "Misc Amount (deliveries)", "Commission Amt",
    "Misc Reimburse (grocery runs)", "2% S-Corp Medical",
]
SALARIED_TEMPLATE_ROWS = [
    {"Name": "Mike", "Department": "M"},
    {"Name": "Rod", "Department": "M"},
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ADP_ACCOUNTING_FORMAT = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
ADP_HEADER_FONT = Font(name="Aptos", size=11, bold=True)
ADP_DATA_FONT = Font(name="Aptos", size=11, bold=False)
ADP_THIN_BORDER = Border(*(Side(style="thin"),) * 4)
ADP_NA_FILL = PatternFill("solid", fgColor="000000")
ADP_TEXT_COLUMNS = {"Name", "Department"}
TOTALS_FILL = PatternFill("solid", fgColor="D9D9D9")
ADP_COLUMN_WIDTHS = {
    "Name": 21.5, "EE Purchases": 21.5, "Department": 15.5,
    "Regular Hours": 15.5, "Salary Amount": 16.2, "Overtime Hours": 15.5,
    "CC Tips Owed": 16.5, "NQ Overtime": 24.2, "NQ Tips": 11.0,
    "PTO Hours": 11.0, "Bonus Amount": 16.5, "Misc Amount (deliveries)": 18.5,
    "Commission Amt": 18.5, "Misc Reimburse (grocery runs)": 22.3,
    "2% S-Corp Medical": 16.7,
}

FONT_NAME = "Aptos"
INPUT_FONT = Font(name=FONT_NAME, size=11, color="0000FF")
FORMULA_FONT = Font(name=FONT_NAME, size=11, color="000000")
BOLD_FONT = Font(name=FONT_NAME, size=11, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=13, bold=True)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")
THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


class ReportError(Exception):
    """Raised for problems the admin needs to see (bad file, no data, etc.)."""


# ---------------------------------------------------------------------------
# Thao's automatic PTO makeup for short weeks (ported from timeclock_to_adp.py)
# ---------------------------------------------------------------------------
THAO_NAME = "Thao Nguyen"


def thao_pto_makeup_hours(records, thao_name, weekly_target=OT_THRESHOLD_HOURS):
    """Per Rod: Thao is expected to work ~40 hrs/week, and if a workweek in
    this pay period comes up short, the difference is made up out of her
    PTO balance. Evaluated independently PER WORKWEEK -- a week that runs
    over 40 does NOT offset a different week's shortfall (confirmed with
    Rod: 45 hrs week 1 / 30 hrs week 2 -> 0 PTO for week 1, 10 PTO for
    week 2, not a single "80 hrs for the period" check).

    Only counts a week where she has SOME clocked hours but less than the
    target -- a week with zero punches at all is assumed to already be an
    actual approved time-off request (a manually-logged PTO request in
    Firestore), so this doesn't double-add PTO on top of that. Uses the
    same paid-hours basis (punches + paid-break minutes on days with >2
    punches) as the Regular/OT split above, so "short" here means the
    same thing it means on her Regular Hours cell.

    Returns (total_makeup_hours, [(week_start_date, hours_worked,
    hours_added), ...]) -- the list only has entries for weeks that were
    actually short. `records` should be the post-sister-company-folding
    records, so a week where her hours are combined with an alias is
    evaluated on her true combined total.
    """
    all_dates = sorted({d for _, d, _ in records})
    if not all_dates:
        return 0.0, []
    pay_period_weeks = sorted({week_start(d) for d in all_dates})

    daily_hours = defaultdict(float)
    daily_count = defaultdict(int)
    for name, date, amt in records:
        if name != thao_name:
            continue
        daily_hours[date] += amt
        daily_count[date] += 1

    week_worked = defaultdict(float)
    for date, hrs in daily_hours.items():
        paid = hrs
        if daily_count[date] > 2:
            paid += break_minutes_for_hours(hrs) / 60.0
        week_worked[week_start(date)] += paid

    makeup = 0.0
    shortfalls = []
    for wk in pay_period_weeks:
        worked = round(week_worked.get(wk, 0.0), 4)
        if 0 < worked < weekly_target:
            added = round(weekly_target - worked, 2)
            makeup += added
            shortfalls.append((wk, worked, added))
    return round(makeup, 2), shortfalls


# ---------------------------------------------------------------------------
# Rod's 2% S-Corp Medical (ported from timeclock_to_adp.py)
# ---------------------------------------------------------------------------
ROD_ROW_NAME = "Rod"
# A confirmed real ADP pay date (a Friday). Every biweekly pay date is
# exactly a multiple of 14 days from this one -- used to tell which
# calendar-month pay period a run is (1st/2nd/3rd) for the rule below.
PAY_DATE_ANCHOR = datetime(2026, 7, 17)
ROD_MEDICAL_AMOUNT = 677.12       # every pay period...
ROD_MEDICAL_THIRD_PERIOD = 0.0    # ...EXCEPT the 3rd biweekly pay date in a calendar month


def is_pay_date(d):
    return (d.date() - PAY_DATE_ANCHOR.date()).days % 14 == 0


def nth_pay_period_of_month(pay_date):
    """Which biweekly pay date this is within its calendar month (1, 2, or
    3 -- most months get 2, some get 3, under the fixed 14-day cadence).
    Returns None if pay_date itself doesn't line up with that cadence at
    all (shouldn't happen for a pay date picked from the app's dropdown)."""
    first_of_month = pay_date.replace(day=1)
    next_month = (datetime(first_of_month.year + 1, 1, 1) if first_of_month.month == 12
                  else datetime(first_of_month.year, first_of_month.month + 1, 1))
    count = 0
    found = None
    d = first_of_month
    while d < next_month:
        if is_pay_date(d):
            count += 1
            if d.date() == pay_date.date():
                found = count
        d += timedelta(days=1)
    return found


def rod_medical_amount(pay_date):
    """Returns (amount, nth) -- nth is None if pay_date doesn't match the
    known biweekly cadence, in which case the caller should leave the
    amount blank for manual entry rather than guess."""
    nth = nth_pay_period_of_month(pay_date)
    if nth is None:
        return None, None
    return (ROD_MEDICAL_THIRD_PERIOD if nth == 3 else ROD_MEDICAL_AMOUNT), nth


# ---------------------------------------------------------------------------
# Sister-company (e.g. Easy Entrées) hours folding + earnout pivot
# (ported from timeclock_to_adp.py)
# ---------------------------------------------------------------------------
def apply_entity_map(records, sister_map):
    """Returns mapped_records: the same (name, date, amt) records, but with
    any sister-company alias name replaced by its canonical employee name
    -- this is what gets fed into compute_hours_and_wages so OT, breaks,
    man-days, tips and wages are all calculated on the person's COMBINED
    hours (one employee = one FLSA hours count). Use build_earnout_pivot
    on the ORIGINAL (pre-fold) records to get the alias-vs-canonical
    breakdown for the earnout schedule -- nothing here affects that."""
    if not sister_map:
        return list(records)
    mapped = []
    for name, date, amt in records:
        alias = sister_map.get(name)
        mapped.append((alias["canonical"], date, amt) if alias else (name, date, amt))
    return mapped


def build_earnout_pivot(records, sister_map):
    """Builds a pivot table exactly like Rod's reference: raw timeclock
    names as rows (the alias AND the real employee's own name, shown
    separately -- not combined), every date in the pay period as columns
    (blank where that name has no hours that day), a "Grand Total" row/
    column, and a rightmost % column showing each row's share of that
    person's combined total. One such table per employee who has a
    sister-company split; grouped by entity label.

    `records` must be the ORIGINAL raw (name, date, amt) records --
    before alias names get remapped to their canonical employee -- so
    each name's own row can still be shown.

    Returns dict entity_label -> dict canonical_name -> {
        "dates": [sorted date list, the full pay-period range],
        "rows": [(raw_name, [hours or None per date], row_total)],
        "totals_per_date": [combined hours per date, None if none],
        "grand_total": combined total hours for this person,
    }
    """
    if not sister_map:
        return {}

    all_dates = sorted({d for _, d, _ in records})
    raw_daily = defaultdict(float)
    raw_seen = defaultdict(bool)
    for name, date, amt in records:
        raw_daily[(name, date)] += amt
        raw_seen[(name, date)] = True

    groups = defaultdict(lambda: defaultdict(set))  # label -> canonical -> {alias names}
    for alias, info in sister_map.items():
        groups[info["entity_label"]][info["canonical"]].add(alias)

    result = {}
    for label, by_canonical in groups.items():
        result[label] = {}
        for canonical, aliases in by_canonical.items():
            raw_names = sorted(aliases | {canonical})
            rows = []
            totals_per_date = [0.0] * len(all_dates)
            date_has_data = [False] * len(all_dates)
            for raw_name in raw_names:
                values = []
                row_total = 0.0
                for i, date in enumerate(all_dates):
                    if raw_seen.get((raw_name, date)):
                        h = raw_daily[(raw_name, date)]
                        values.append(round(h, 2))
                        totals_per_date[i] += h
                        date_has_data[i] = True
                        row_total += h
                    else:
                        values.append(None)
                rows.append((raw_name, values, round(row_total, 2)))
            grand_total = round(sum(row_total for _, _, row_total in rows), 2)
            result[label][canonical] = {
                "dates": all_dates,
                "rows": rows,
                "totals_per_date": [round(v, 2) if date_has_data[i] else None
                                    for i, v in enumerate(totals_per_date)],
                "grand_total": grand_total,
            }
    return result


# ---------------------------------------------------------------------------
# Step 1: parse the raw timeclock export (from decoded text, not a path)
# ---------------------------------------------------------------------------
def parse_raw_export_text(text):
    records = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if len(row) <= CSV_COL_TRANS_TTL:
            continue
        name = row[CSV_COL_NAME].strip()
        date_raw = row[CSV_COL_DATE].strip()
        amt_raw = row[CSV_COL_TRANS_TTL].strip()
        if not name or not date_raw:
            continue
        try:
            date = datetime.strptime(date_raw, CSV_DATE_FORMAT)
        except ValueError:
            continue
        try:
            amt = float(amt_raw.replace(",", "")) if amt_raw not in ("", "-") else 0.0
        except ValueError:
            amt = 0.0
        records.append((name, date, amt))
    if not records:
        raise ReportError(
            "No transaction rows found in the uploaded CSV. Is this the raw "
            "'Transaction Report' export from the timeclock system?"
        )
    return records


def break_minutes_for_hours(hours):
    minutes = 0
    for threshold, m in BREAKS_TABLE:
        if hours >= threshold:
            minutes = m
        else:
            break
    return minutes


def week_start(date):
    offset = (date.weekday() - WORKWEEK_START_WEEKDAY) % 7
    return date - timedelta(days=offset)


def man_days_for_hours(hours):
    if hours > MAN_DAY_FULL_THRESHOLD:
        return MAN_DAY_FULL_CREDIT
    if hours > 0:
        return MAN_DAY_HALF_CREDIT
    return 0.0


def compute_hours_and_wages(records, employees):
    daily_hours = defaultdict(float)
    daily_count = defaultdict(int)
    names_seen = set()
    for name, date, amt in records:
        daily_hours[(name, date)] += amt
        daily_count[(name, date)] += 1
        names_seen.add(name)

    unknown = sorted(n for n in names_seen
                      if n not in employees or employees[n]["rate"] is None)

    all_dates = sorted({d for _, d in daily_hours.keys()})
    results = {}
    for name, info in employees.items():
        if name not in names_seen or info["rate"] is None:
            continue
        week_punched = defaultdict(float)
        week_break = defaultdict(float)
        for date in all_dates:
            key = (name, date)
            if key not in daily_hours:
                continue
            wk = week_start(date)
            week_punched[wk] += daily_hours[key]
            if daily_count[key] > 2:
                week_break[wk] += break_minutes_for_hours(daily_hours[key]) / 60.0

        regular_total = ot_total = wages_total = 0.0
        rate = info["rate"]
        for wk in week_punched:
            paid_basis = week_punched[wk] + week_break[wk]
            reg = min(paid_basis, OT_THRESHOLD_HOURS)
            ot = max(paid_basis - OT_THRESHOLD_HOURS, 0.0)
            regular_total += reg
            ot_total += ot
            wages_total += rate * reg + OT_MULTIPLIER * rate * ot

        man_days_total = sum(
            man_days_for_hours(daily_hours[(name, date)])
            for date in all_dates if (name, date) in daily_hours
        )
        results[name] = {
            "regular_hours": round(regular_total, 4),
            "ot_hours": round(ot_total, 4),
            "wages": round(wages_total, 2),
            "man_days": round(man_days_total, 2),
        }
    return results, unknown, all_dates


def allocate_tips_with_drivers(computed, employees, net_pool, driver_days):
    """driver_days: dict driver_name -> total days driven (flat 1.0/day, no
    half-day rule). Returns (w2_tips: {name: $}, driver_tips: {name: $}),
    both dicts only containing people who actually share in the pool."""
    w2_eligible = [
        name for name, info in employees.items()
        if info["tip_eligible"] and name in computed and computed[name]["man_days"] > 0
    ]
    driver_eligible = [n for n, d in driver_days.items() if d and d > 0]

    total_days = sum(computed[n]["man_days"] for n in w2_eligible) + \
        sum(driver_days[n] for n in driver_eligible)
    if total_days <= 0:
        return {}, {}
    rate_per_day = net_pool / total_days
    w2_tips = {n: round(rate_per_day * computed[n]["man_days"], 2) for n in w2_eligible}
    driver_tips = {n: round(rate_per_day * driver_days[n], 2) for n in driver_eligible}
    return w2_tips, driver_tips


# ---------------------------------------------------------------------------
# Workbook build
# ---------------------------------------------------------------------------
def title_block(ws, title, source_lines):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    r = 2
    for line in source_lines:
        ws.cell(row=r, column=1, value=line).font = NOTE_FONT
        r += 1
    return r + 1


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, min_width=9, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            L = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, min_width), min(L + 2, max_width))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_report(csv_text, employees, order, pay_date, raw_csv_name,
                  total_tip_revenue, bonnie_brae, swift, driver_info,
                  pending_pto, pending_purchases, pending_misc_amt,
                  pending_misc_reimb, sister_map=None):
    """employees/order: same shape as before (dict name -> {department,
    rate, tip_eligible}, plus the name order list).

    driver_info: dict driver_name -> {"days": <days driven, flat 1.0/day
    credit>, "deliveries": <$>, "setups": <$>} -- entered by Larry (or a
    Manager backing him up) on the Tip Pool tab. "days" feeds the
    tip-pool payout (same as before); deliveries/setups feed the
    "Driver Payroll (1099s) Recap" sheet. That sheet's Tips column is NOT
    entered by Larry -- it's the same computed tip-pool payout as the
    "Driver Tip Payouts (1099s)" sheet (allocate_tips_with_drivers), since
    a driver's Tips figure only exists once the tip pool math runs. This
    replaces the separate "Driver Payroll and EE Tips" workbook Larry used
    to maintain by hand.

    sister_map: dict alias_name -> {"canonical": <real employee name>,
    "entity_label": <e.g. "Easy Entrées">} from the `sisterCompanyAliases`
    Firestore collection (Owner-maintained). Hours punched under an alias
    are folded into the canonical employee's combined hours for every
    wage/OT/break/man-days/tips calculation, and also broken out on a
    separate "Earnout - <label>" sheet per entity for reference. Pass
    None/empty if there are no aliases -- nothing changes.

    pending_* : lists of dicts, each {"id": <firestore doc id>,
    "employeeName": ..., "hours"/"amount": ..., ...} -- every currently
    unprocessed record from that Firestore collection, handed in by
    main.py. This function does not touch Firestore itself; it just
    aggregates per employee for the workbook and tells the caller (via
    consumed_ids in the return value) which record ids it used, so
    main.py can stamp them processed -- but only if the caller chooses
    to finalize (see main.py).

    Returns (workbook_bytes, summary, warnings, consumed_ids) where
    consumed_ids is {"ptoRequests": [id, ...], "employeePurchases": [...],
    "miscAmounts": [...], "miscReimbursements": [...]}.
    """
    warnings = []
    sister_map = sister_map or {}
    driver_info = driver_info or {}
    driver_days = {name: info.get("days", 0) for name, info in driver_info.items()}

    original_records = parse_raw_export_text(csv_text)
    records = apply_entity_map(original_records, sister_map)
    earnout_pivot = build_earnout_pivot(original_records, sister_map)

    computed, unknown_names, all_dates = compute_hours_and_wages(records, employees)
    if unknown_names:
        warnings.append(
            "These names appear in the timeclock export but have no wage rate "
            "on file, so their hours are not in this report: " + ", ".join(unknown_names)
        )

    thao_makeup, thao_shortfalls = (0.0, [])
    if THAO_NAME in employees:
        thao_makeup, thao_shortfalls = thao_pto_makeup_hours(records, THAO_NAME)
        if thao_makeup:
            weeks_desc = ", ".join(
                f"week of {wk:%m/%d} (worked {worked:.2f} hrs, +{added:.2f} PTO)"
                for wk, worked, added in thao_shortfalls
            )
            warnings.append(
                f"Added {thao_makeup:.2f} automatic PTO makeup hour(s) for "
                f"{THAO_NAME} for week(s) under 40 hrs worked: {weeks_desc}."
            )

    rod_medical, rod_medical_nth = rod_medical_amount(pay_date)
    if rod_medical is None:
        warnings.append(
            f"{pay_date:%m/%d/%Y} doesn't line up with the known biweekly pay "
            "cadence, so Rod's 2% S-Corp Medical was left blank for manual entry."
        )

    net_pool = round(total_tip_revenue - bonnie_brae - swift, 2)
    w2_tips, driver_tips = allocate_tips_with_drivers(computed, employees, net_pool, driver_days)

    def sum_by_employee(records_list, amount_field, known_names):
        totals = defaultdict(float)
        consumed_ids = []
        unresolved = []
        for r in records_list:
            name = (r.get("employeeName") or "").strip()
            amt = r.get(amount_field)
            if not name or not isinstance(amt, (int, float)):
                continue
            if name not in known_names:
                unresolved.append(name)
                continue
            totals[name] += amt
            consumed_ids.append(r["id"])
        return dict(totals), consumed_ids, unresolved

    pto_hours, pto_ids, u1 = sum_by_employee(pending_pto, "hours", employees)
    ee_purchases, purch_ids, u2 = sum_by_employee(pending_purchases, "amount", employees)
    misc_amt, misc_amt_ids, u3 = sum_by_employee(pending_misc_amt, "amount", employees)
    misc_reimb, misc_reimb_ids, u4 = sum_by_employee(pending_misc_reimb, "amount", employees)
    if thao_makeup:
        pto_hours[THAO_NAME] = round(pto_hours.get(THAO_NAME, 0.0) + thao_makeup, 2)
    unresolved_names = sorted(set(u1) | set(u2) | set(u3) | set(u4))
    if unresolved_names:
        warnings.append(
            "These pending requests reference a name that isn't in the "
            "employee list, so they're NOT included: " + ", ".join(unresolved_names)
        )
    consumed_ids = {
        "ptoRequests": pto_ids, "employeePurchases": purch_ids,
        "miscAmounts": misc_amt_ids, "miscReimbursements": misc_reimb_ids,
    }

    # -------- daily records for Sheet 1 --------
    daily_hours = defaultdict(float)
    daily_count = defaultdict(int)
    for name, date, amt in records:
        daily_hours[(name, date)] += amt
        daily_count[(name, date)] += 1
    names_seen = sorted({n for n, _, _ in records})
    daily_rows = []
    for name in names_seen:
        for d in all_dates:
            key = (name, d)
            if key in daily_hours:
                hrs = daily_hours[key]
                cnt = daily_count[key]
                brk = break_minutes_for_hours(hrs) / 60.0 if cnt > 2 else 0.0
                daily_rows.append({
                    "name": name, "date": d, "weekday": d.strftime("%a"),
                    "hours": round(hrs, 4), "punches": cnt, "break_hrs": round(brk, 4),
                    "man_day": man_days_for_hours(hrs), "week_start": week_start(d),
                })

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- Sheet 0: Employee Config ----
    ws0 = wb.create_sheet("0 - Employee Config")
    r0 = title_block(ws0, "Employee Config", [
        "Source: the 'employees' collection (Firestore), maintained by the Owner.",
        "Blue = source input. Every other sheet looks up rate/department/tip-eligibility from here.",
    ])
    for i, h in enumerate(["Employee", "Department", "Hourly Rate", "Tip Eligible?"], start=1):
        ws0.cell(row=r0, column=i, value=h)
    style_header(ws0, r0, 4)
    cfg_first_row = r0 + 1
    for i, name in enumerate(order):
        info = employees[name]
        row = cfg_first_row + i
        ws0.cell(row=row, column=1, value=name).font = INPUT_FONT
        ws0.cell(row=row, column=2, value=info["department"]).font = INPUT_FONT
        c = ws0.cell(row=row, column=3, value=info["rate"])
        c.font = INPUT_FONT
        c.number_format = "$#,##0.00"
        ws0.cell(row=row, column=4, value="y" if info["tip_eligible"] else "n").font = INPUT_FONT
        for cc in range(1, 5):
            ws0.cell(row=row, column=cc).border = THIN
    cfg_last_row = cfg_first_row + len(order) - 1
    autosize(ws0)
    CFG_SHEET = "'0 - Employee Config'"

    # ---- Sheet 1: Daily Punches ----
    ws1 = wb.create_sheet("1 - Daily Punches")
    r1 = title_block(ws1, "Daily Punches", [
        f"Source: {raw_csv_name} (raw timeclock export, uploaded for this report).",
        "Break Hrs Credited = paid rest-break minutes, only on a day with more than 2 punches (Colorado law).",
        "Man-Day Credit = 1.0 if worked >4 hrs that day, 0.5 if worked some but <=4 hrs, 0 if didn't work.",
    ])
    headers1 = ["Employee", "Date", "Day", "Punches That Day", "Hours Worked",
                "Break Hrs Credited", "Paid Hours (Worked+Break)", "Man-Day Credit", "Week Start (Sat)"]
    for i, h in enumerate(headers1, start=1):
        ws1.cell(row=r1, column=i, value=h)
    style_header(ws1, r1, len(headers1))
    d_first_row = r1 + 1
    daily_sorted = sorted(daily_rows, key=lambda x: (x["name"], x["date"]))
    for i, rec in enumerate(daily_sorted):
        row = d_first_row + i
        ws1.cell(row=row, column=1, value=rec["name"]).font = INPUT_FONT
        dcell = ws1.cell(row=row, column=2, value=rec["date"]); dcell.number_format = "mm/dd/yyyy"; dcell.font = INPUT_FONT
        ws1.cell(row=row, column=3, value=rec["weekday"]).font = INPUT_FONT
        ws1.cell(row=row, column=4, value=rec["punches"]).font = INPUT_FONT
        ws1.cell(row=row, column=5, value=rec["hours"]).font = INPUT_FONT
        ws1.cell(row=row, column=6, value=rec["break_hrs"]).font = INPUT_FONT
        ws1.cell(row=row, column=7, value=f"=E{row}+F{row}").font = FORMULA_FONT
        ws1.cell(row=row, column=8, value=rec["man_day"]).font = INPUT_FONT
        wcell = ws1.cell(row=row, column=9, value=rec["week_start"]); wcell.number_format = "mm/dd/yyyy"; wcell.font = INPUT_FONT
        for cc in range(1, 10):
            ws1.cell(row=row, column=cc).border = THIN
    d_last_row = d_first_row + len(daily_sorted) - 1 if daily_sorted else d_first_row
    autosize(ws1)
    DAILY_SHEET = "'1 - Daily Punches'"

    # ---- Sheet 2: Weekly Hours & Wages ----
    ws2 = wb.create_sheet("2 - Weekly Hours & Wages")
    r2 = title_block(ws2, "Weekly Hours & Wages", [
        f"Paid Hours pulled from {DAILY_SHEET}. Rate pulled from {CFG_SHEET}.",
        "Workweek = Saturday-Friday, 40 hr/week overtime threshold, 1.5x OT multiplier.",
        "Period Totals block below sums every week in this upload -- these feed the ADP Entry sheet.",
    ])
    headers2 = ["Employee", "Week Start", "Week End", "Paid Hours", "Regular Hours",
                "Overtime Hours", "Hourly Rate", "Regular Wages", "OT Wages", "Gross Wages"]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=r2, column=i, value=h)
    style_header(ws2, r2, len(headers2))
    week_starts = sorted({rec["week_start"] for rec in daily_rows})
    wk_first_row = r2 + 1
    row = wk_first_row
    names_with_hours = {d["name"] for d in daily_rows}
    for name in order:
        if name not in names_with_hours:
            continue
        for wk in week_starts:
            wk_end = wk + timedelta(days=6)
            ws2.cell(row=row, column=1, value=name).font = INPUT_FONT
            c2 = ws2.cell(row=row, column=2, value=wk); c2.number_format = "mm/dd/yyyy"; c2.font = INPUT_FONT
            c3 = ws2.cell(row=row, column=3, value=wk_end); c3.number_format = "mm/dd/yyyy"; c3.font = INPUT_FONT
            paid_f = (f'=SUMIFS({DAILY_SHEET}!$G${d_first_row}:$G${d_last_row},'
                      f'{DAILY_SHEET}!$A${d_first_row}:$A${d_last_row},A{row},'
                      f'{DAILY_SHEET}!$I${d_first_row}:$I${d_last_row},B{row})')
            ws2.cell(row=row, column=4, value=paid_f).font = FORMULA_FONT
            ws2.cell(row=row, column=5, value=f"=MIN(D{row},40)").font = FORMULA_FONT
            ws2.cell(row=row, column=6, value=f"=MAX(D{row}-40,0)").font = FORMULA_FONT
            rate_f = f'=VLOOKUP(A{row},{CFG_SHEET}!$A${cfg_first_row}:$C${cfg_last_row},3,FALSE)'
            ws2.cell(row=row, column=7, value=rate_f).font = FORMULA_FONT
            ws2.cell(row=row, column=7).number_format = "$#,##0.00"
            ws2.cell(row=row, column=8, value=f"=E{row}*G{row}").font = FORMULA_FONT
            ws2.cell(row=row, column=9, value=f"=F{row}*G{row}*1.5").font = FORMULA_FONT
            ws2.cell(row=row, column=10, value=f"=H{row}+I{row}").font = FORMULA_FONT
            for cc in (4, 5, 6, 8, 9, 10):
                ws2.cell(row=row, column=cc).number_format = "#,##0.00"
            for cc in range(1, 11):
                ws2.cell(row=row, column=cc).border = THIN
            row += 1
    wk_last_row = max(row - 1, wk_first_row)

    row += 2
    ws2.cell(row=row, column=1, value="Period Totals (feeds ADP Entry)").font = BOLD_FONT
    row += 1
    headers2b = ["Employee", "Total Regular Hours", "Total OT Hours", "Total Gross Wages"]
    for i, h in enumerate(headers2b, start=1):
        ws2.cell(row=row, column=i, value=h)
    style_header(ws2, row, len(headers2b))
    pt_first_row = row + 1
    row += 1
    for name in order:
        if name not in names_with_hours:
            continue
        ws2.cell(row=row, column=1, value=name).font = INPUT_FONT
        ws2.cell(row=row, column=2, value=f"=SUMIF($A${wk_first_row}:$A${wk_last_row},A{row},$E${wk_first_row}:$E${wk_last_row})").font = FORMULA_FONT
        ws2.cell(row=row, column=3, value=f"=SUMIF($A${wk_first_row}:$A${wk_last_row},A{row},$F${wk_first_row}:$F${wk_last_row})").font = FORMULA_FONT
        ws2.cell(row=row, column=4, value=f"=SUMIF($A${wk_first_row}:$A${wk_last_row},A{row},$J${wk_first_row}:$J${wk_last_row})").font = FORMULA_FONT
        for cc in (2, 3, 4):
            ws2.cell(row=row, column=cc).number_format = "#,##0.00"
        for cc in range(1, 5):
            ws2.cell(row=row, column=cc).border = THIN
        row += 1
    pt_last_row = max(row - 1, pt_first_row)
    autosize(ws2)
    WEEKLY_SHEET = "'2 - Weekly Hours & Wages'"

    # ---- Sheet 3: Man-Days & Tips Calc ----
    ws3 = wb.create_sheet("3 - Man-Days & Tips Calc")
    r3 = title_block(ws3, "Man-Days & Tips Calculation", [
        "Total Tip Revenue / Bonnie Brae / Swift: entered on the admin report page for this period.",
        "SIMPLIFIED MODEL: days worked are computed from THIS UPLOAD ONLY (one pay period's raw timeclock",
        "export), not the fully-proven 4-week/2-pay-period rolling window (that model is proven correct but",
        "needs a persisted man-days history across periods, not yet built -- ask Rod/Claude to add it).",
        "1099 driver day-counts come from the tips web form (Larry, or a Manager filling in for him). Drivers",
        "get a flat 1.0 credit per day driven -- never the 0.5-short-day rule that applies to W2 employees.",
        "Share = a person's Days / everyone's combined Days (W2 tip-eligible + all drivers, same pool).",
        "Tip $ = Share x Net Pool. Sam Gray is excluded from any share regardless of days worked.",
    ])
    ws3.cell(row=r3, column=1, value="Total Tip Revenue").font = BOLD_FONT
    c = ws3.cell(row=r3, column=2, value=total_tip_revenue); c.font = INPUT_FONT; c.number_format = "$#,##0.00"
    ws3.cell(row=r3 + 1, column=1, value="Less: Bonnie Brae").font = BOLD_FONT
    c = ws3.cell(row=r3 + 1, column=2, value=bonnie_brae); c.font = INPUT_FONT; c.number_format = "$#,##0.00"
    ws3.cell(row=r3 + 2, column=1, value="Less: Swift").font = BOLD_FONT
    c = ws3.cell(row=r3 + 2, column=2, value=swift); c.font = INPUT_FONT; c.number_format = "$#,##0.00"
    ws3.cell(row=r3 + 3, column=1, value="Net Tip Pool").font = BOLD_FONT
    netpool_cell = f"B{r3 + 3}"
    c = ws3.cell(row=r3 + 3, column=2, value=f"=B{r3}-B{r3+1}-B{r3+2}"); c.font = FORMULA_FONT; c.number_format = "$#,##0.00"
    for rr in range(r3, r3 + 4):
        for cc in (1, 2):
            ws3.cell(row=rr, column=cc).border = THIN

    row = r3 + 5
    ws3.cell(row=row, column=1, value="W2 Employees").font = BOLD_FONT
    row += 1
    headers3 = ["Employee", "Days Worked (this upload)", "Tip Eligible?", "Share of Pool", "Tip $ Payout"]
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=row, column=i, value=h)
    style_header(ws3, row, len(headers3))
    row += 1
    w2_rows_start = row
    for name in order:
        ws3.cell(row=row, column=1, value=name).font = INPUT_FONT
        days_f = (f'=SUMIF({DAILY_SHEET}!$A${d_first_row}:$A${d_last_row},A{row},'
                  f'{DAILY_SHEET}!$H${d_first_row}:$H${d_last_row})')
        ws3.cell(row=row, column=2, value=days_f).font = FORMULA_FONT
        elig_f = f'=VLOOKUP(A{row},{CFG_SHEET}!$A${cfg_first_row}:$D${cfg_last_row},4,FALSE)'
        ws3.cell(row=row, column=3, value=elig_f).font = FORMULA_FONT
        for cc in range(1, 6):
            ws3.cell(row=row, column=cc).border = THIN
        row += 1
    w2_rows_end = row - 1

    row += 1
    ws3.cell(row=row, column=1, value="1099 Drivers (always 1.0 day credit/day driven)").font = BOLD_FONT
    row += 1
    headers3b = ["Driver", "Days Driven", "Tip Eligible?", "Share of Pool", "Tip $ Payout"]
    for i, h in enumerate(headers3b, start=1):
        ws3.cell(row=row, column=i, value=h)
    style_header(ws3, row, len(headers3b))
    row += 1
    drv_rows_start = row
    for dname, ddays in driver_days.items():
        ws3.cell(row=row, column=1, value=dname).font = INPUT_FONT
        ws3.cell(row=row, column=2, value=ddays).font = INPUT_FONT
        ws3.cell(row=row, column=3, value="y").font = INPUT_FONT
        for cc in range(1, 6):
            ws3.cell(row=row, column=cc).border = THIN
        row += 1
    drv_rows_end = row - 1 if driver_days else row

    row += 1
    ws3.cell(row=row, column=1, value="Grand Total Days (tip-eligible only -- denominator for Share)").font = BOLD_FONT
    if driver_days:
        grand_total_f = (f'=SUMIF(C{w2_rows_start}:C{w2_rows_end},"y",B{w2_rows_start}:B{w2_rows_end})'
                          f'+SUM(B{drv_rows_start}:B{drv_rows_end})')
    else:
        grand_total_f = f'=SUMIF(C{w2_rows_start}:C{w2_rows_end},"y",B{w2_rows_start}:B{w2_rows_end})'
    gt_cell = ws3.cell(row=row, column=2, value=grand_total_f)
    gt_cell.font = FORMULA_FONT
    gt_cell_ref = f"$B${row}"
    for cc in (1, 2):
        ws3.cell(row=row, column=cc).border = THIN

    for rr in list(range(w2_rows_start, w2_rows_end + 1)) + \
            (list(range(drv_rows_start, drv_rows_end + 1)) if driver_days else []):
        share_f = f'=IF(C{rr}="y",B{rr}/{gt_cell_ref},0)'
        ws3.cell(row=rr, column=4, value=share_f).font = FORMULA_FONT
        ws3.cell(row=rr, column=4).number_format = "0.0000%"
        tip_f = f'=D{rr}*{netpool_cell}'
        ws3.cell(row=rr, column=5, value=tip_f).font = FORMULA_FONT
        ws3.cell(row=rr, column=5).number_format = "$#,##0.00"
    autosize(ws3)
    TIPS_SHEET = "'3 - Man-Days & Tips Calc'"

    # ---- Sheet 4: Deductions ----
    ws4 = wb.create_sheet("4 - Deductions")
    r4 = title_block(ws4, "Deductions (PTO / Purchases / Misc / S-Corp)", [
        "Source: pending requests logged in the app by a Manager or the Owner (PTO / Employee",
        "Purchases / Misc Amount / Misc Reimbursement collections), swept up as of this report run.",
    ])
    headers4 = ["Employee", "PTO Hours", "EE Purchases", "Misc Amount (deliveries)",
                "Misc Reimburse (grocery runs)"]
    for i, h in enumerate(headers4, start=1):
        ws4.cell(row=r4, column=i, value=h)
    style_header(ws4, r4, len(headers4))
    ded_first_row = r4 + 1
    row = ded_first_row
    all_ded_names = sorted(set(pto_hours) | set(ee_purchases) | set(misc_amt) | set(misc_reimb))
    for name in all_ded_names:
        ws4.cell(row=row, column=1, value=name).font = INPUT_FONT
        for i, d in enumerate([pto_hours, ee_purchases, misc_amt, misc_reimb], start=2):
            v = d.get(name)
            c = ws4.cell(row=row, column=i, value=v)
            c.font = INPUT_FONT
            if v is not None:
                c.number_format = "#,##0.00"
        for cc in range(1, 6):
            ws4.cell(row=row, column=cc).border = THIN
        row += 1
    ded_last_row = max(row - 1, ded_first_row)
    autosize(ws4)
    DED_SHEET = "'4 - Deductions'"
    DED_COLS = {"PTO Hours": 2, "EE Purchases": 3, "Misc Amount (deliveries)": 4,
                "Misc Reimburse (grocery runs)": 5}

    # ---- Sheet 5 (LAST): ADP Entry ----
    ws5 = wb.create_sheet("ADP Entry")
    ws5.append(ADP_COLUMNS)
    for cell in ws5[1]:
        cell.font = ADP_HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = ADP_THIN_BORDER
    ws5.row_dimensions[1].height = 32
    ws5.freeze_panes = "B1"
    SALARIED_NA_COLUMNS = {"Overtime Hours", "CC Tips Owed", "NQ Overtime", "NQ Tips"}
    HOURLY_NA_COLUMNS = {"Salary Amount"}
    col_index = {c: i + 1 for i, c in enumerate(ADP_COLUMNS)}

    def ded_lookup(field_col, name_cell):
        return (f'=IFERROR(VLOOKUP({name_cell},{DED_SHEET}!$A${ded_first_row}:$E${ded_last_row},'
                f'{field_col},FALSE),"")')

    def write_row(name, na_columns, is_salaried):
        r = ws5.max_row + 1
        ws5.append([None] * len(ADP_COLUMNS))
        ws5.row_dimensions[r].height = 16
        name_cell = f"A{r}"
        for col_name in ADP_COLUMNS:
            c = col_index[col_name]
            cell = ws5.cell(row=r, column=c)
            cell.font = ADP_DATA_FONT
            cell.border = ADP_THIN_BORDER
            cell.number_format = "General" if col_name in ADP_TEXT_COLUMNS else ADP_ACCOUNTING_FORMAT
            if col_name in na_columns:
                cell.fill = ADP_NA_FILL
                continue
            if col_name == "Name":
                cell.value = name
            elif col_name == "Department":
                cell.value = "M" if is_salaried else f'=VLOOKUP({name_cell},{CFG_SHEET}!$A${cfg_first_row}:$D${cfg_last_row},2,FALSE)'
            elif col_name == "Regular Hours" and not is_salaried:
                cell.value = f'=IFERROR(VLOOKUP({name_cell},{WEEKLY_SHEET}!$A${pt_first_row}:$D${pt_last_row},2,FALSE),0)'
            elif col_name == "Overtime Hours" and not is_salaried:
                cell.value = f'=IFERROR(VLOOKUP({name_cell},{WEEKLY_SHEET}!$A${pt_first_row}:$D${pt_last_row},3,FALSE),0)'
            elif col_name == "CC Tips Owed" and not is_salaried:
                cell.value = f'=IFERROR(VLOOKUP({name_cell},{TIPS_SHEET}!$A${w2_rows_start}:$E${w2_rows_end},5,FALSE),"")'
            elif col_name == "2% S-Corp Medical" and name == ROD_ROW_NAME and rod_medical is not None:
                cell.value = rod_medical
            elif col_name in DED_COLS:
                cell.value = ded_lookup(DED_COLS[col_name], name_cell)
        return r

    for template in SALARIED_TEMPLATE_ROWS:
        write_row(template["Name"], SALARIED_NA_COLUMNS, True)
    for name in order:
        write_row(name, HOURLY_NA_COLUMNS, False)

    first_data_row = 2
    last_data_row = ws5.max_row
    total_row = last_data_row + 1
    ws5.append([None] * len(ADP_COLUMNS))
    ws5.row_dimensions[total_row].height = 18
    for col_name in ADP_COLUMNS:
        cell = ws5.cell(row=total_row, column=col_index[col_name])
        cell.font = ADP_HEADER_FONT
        cell.border = ADP_THIN_BORDER
        cell.fill = TOTALS_FILL
        if col_name == "Name":
            cell.value = "Totals"
            cell.number_format = "General"
        elif col_name in ADP_TEXT_COLUMNS:
            cell.number_format = "General"
        else:
            col_letter = get_column_letter(col_index[col_name])
            cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            cell.number_format = ADP_ACCOUNTING_FORMAT
    for col_name, width in ADP_COLUMN_WIDTHS.items():
        ws5.column_dimensions[get_column_letter(col_index[col_name])].width = width

    if driver_days:
        ws6 = wb.create_sheet("Driver Tip Payouts (1099s)")
        r6 = title_block(ws6, "Driver Tip Payouts", [
            "Not part of ADP Entry -- 1099 contractors aren't payroll. Pulled from " + TIPS_SHEET + ".",
        ])
        for i, h in enumerate(["Driver", "Days Driven", "Tip $ Payout"], start=1):
            ws6.cell(row=r6, column=i, value=h)
        style_header(ws6, r6, 3)
        row = r6 + 1
        for dname in driver_days:
            ws6.cell(row=row, column=1, value=dname).font = INPUT_FONT
            tips_f = f'=VLOOKUP(A{row},{TIPS_SHEET}!$A${drv_rows_start}:$E${drv_rows_end},5,FALSE)'
            ws6.cell(row=row, column=2, value=driver_days[dname]).font = INPUT_FONT
            ws6.cell(row=row, column=3, value=tips_f).font = FORMULA_FONT
            ws6.cell(row=row, column=3).number_format = "$#,##0.00"
            for cc in range(1, 4):
                ws6.cell(row=row, column=cc).border = THIN
            row += 1
        autosize(ws6)

    if driver_info:
        ws7 = wb.create_sheet("Driver Payroll (1099s) Recap")
        r7 = title_block(ws7, "Driver Payroll (1099s) Recap", [
            "Deliveries $ Total and Setups $ Total are NOT counts -- different deliveries/setups pay",
            "different amounts, so each is the driver's own already-added-up dollar total for the period",
            "(entered on the Tip Pool tab, same as the old Driver Payroll workbook).",
            "Tips is NOT entered by Larry -- it's pulled from " + TIPS_SHEET + " (the same tip-pool payout",
            "shown on the Driver Tip Payouts sheet), since that figure only exists once the tip-pool math runs.",
            "Not part of ADP Entry (1099s aren't payroll).",
            "Total = Tips + Deliveries $ Total + Setups $ Total (computed here, not separately entered).",
        ])
        for i, h in enumerate(["Driver", "Tips", "Deliveries $ Total", "Setups $ Total", "Total"], start=1):
            ws7.cell(row=r7, column=i, value=h)
        style_header(ws7, r7, 5)
        row = r7 + 1
        recap_first_row = row
        for dname, info in driver_info.items():
            ws7.cell(row=row, column=1, value=dname).font = INPUT_FONT
            tips_f = f'=VLOOKUP(A{row},{TIPS_SHEET}!$A${drv_rows_start}:$E${drv_rows_end},5,FALSE)'
            c = ws7.cell(row=row, column=2, value=tips_f)
            c.font = FORMULA_FONT
            c.number_format = "$#,##0.00"
            for cc, field in ((3, "deliveries"), (4, "setups")):
                c = ws7.cell(row=row, column=cc, value=info.get(field, 0) or 0)
                c.font = INPUT_FONT
                c.number_format = "$#,##0.00"
            total_f = f"=B{row}+C{row}+D{row}"
            c = ws7.cell(row=row, column=5, value=total_f)
            c.font = FORMULA_FONT
            c.number_format = "$#,##0.00"
            for cc in range(1, 6):
                ws7.cell(row=row, column=cc).border = THIN
            row += 1
        recap_last_row = row - 1
        ws7.cell(row=row, column=1, value="Totals").font = BOLD_FONT
        for cc in (2, 3, 4, 5):
            col_letter = get_column_letter(cc)
            c = ws7.cell(row=row, column=cc,
                         value=f"=SUM({col_letter}{recap_first_row}:{col_letter}{recap_last_row})")
            c.font = BOLD_FONT
            c.number_format = "$#,##0.00"
        for cc in range(1, 6):
            ws7.cell(row=row, column=cc).border = THIN
            ws7.cell(row=row, column=cc).fill = TOTALS_FILL
        autosize(ws7)

    # --- Earnout pivot: mirrors Rod's reference exactly -- a "Sum of amt"
    # style pivot (dates as columns) filtered down to just the alias and
    # real-name rows for each employee with a sister-company split, with a
    # Grand Total row/column and a % column showing each row's share of
    # that employee's combined total. No dollar figure -- Rod's bookkeeper
    # applies payroll burden and splits cost herself from the percentage.
    PCT_FORMAT = "0.0%"
    DATE_FORMAT = "m/d/yyyy"
    for label, by_canonical in (earnout_pivot or {}).items():
        sheet_name = f"Earnout - {label}"[:31]  # Excel sheet name cap
        ws_e = wb.create_sheet(sheet_name)

        for canonical in sorted(by_canonical):
            table = by_canonical[canonical]
            dates = table["dates"]

            ws_e.append(["Sum of amt", "Column Labels"])
            header_row = ws_e.max_row + 1
            ws_e.append(["Row Labels"] + dates + ["Grand Total", "% of total"])
            for c in range(2, 2 + len(dates)):
                ws_e.cell(row=header_row, column=c).number_format = DATE_FORMAT
            style_header(ws_e, header_row, 2 + len(dates))

            for raw_name, values, row_total in table["rows"]:
                ws_e.append([raw_name] + values + [row_total,
                             (row_total / table["grand_total"]) if table["grand_total"] else None])
                ws_e.cell(row=ws_e.max_row, column=2 + len(dates) + 1).number_format = PCT_FORMAT

            ws_e.append(["Grand Total"] + table["totals_per_date"] + [table["grand_total"], None])
            for cell in ws_e[ws_e.max_row]:
                cell.font = BOLD_FONT

            ws_e.append([])  # spacer between employees, if more than one

        for row_cells in ws_e.iter_rows(min_col=2, max_col=1 + len(next(iter(by_canonical.values()))["dates"])):
            for cell in row_cells:
                if isinstance(cell.value, datetime):
                    cell.number_format = DATE_FORMAT

        ws_e.append([])
        ws_e.append([
            "% of total = each row's Grand Total / that employee's combined "
            "Grand Total across both names for the pay period. No payroll-"
            "burden or overtime-premium dollar allocation is calculated "
            "here -- the bookkeeper applies that from these percentages."
        ])
        autosize(ws_e)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    summary = {
        "payDate": pay_date.strftime("%Y-%m-%d"),
        "netTipPool": net_pool,
        "employees": [
            {
                "name": name,
                "department": employees[name]["department"],
                "regularHours": computed.get(name, {}).get("regular_hours", 0),
                "otHours": computed.get(name, {}).get("ot_hours", 0),
                "manDays": computed.get(name, {}).get("man_days", 0),
                "ccTipsOwed": w2_tips.get(name),
                "ptoHours": pto_hours.get(name),
                "eePurchases": ee_purchases.get(name),
                "miscAmount": misc_amt.get(name),
                "miscReimburse": misc_reimb.get(name),
            }
            for name in order if name in names_with_hours
        ],
        "drivers": [
            {
                "name": dname,
                "daysDriven": ddays,
                "tipPayout": driver_tips.get(dname),
                "tips": driver_tips.get(dname, 0),
                "deliveries": driver_info.get(dname, {}).get("deliveries", 0),
                "setups": driver_info.get(dname, {}).get("setups", 0),
            }
            for dname, ddays in driver_days.items()
        ],
        "warnings": warnings,
    }
    return buf.read(), summary, warnings, consumed_ids
